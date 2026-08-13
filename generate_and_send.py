#!/usr/bin/env python3
"""
Task-Tracker → tägliches Update
================================
Liest das Google-Sheet "Task Tracker Malin", nimmt alle Zeilen mit Spalte K = "Y",
lässt daraus von Gemini ein Update im gewohnten Stil formulieren und schickt es
per E-Mail an Malin.

Läuft auf GitHub Actions, Mo–Fr um 17:50 Uhr (Europe/Berlin, sommer-/winterzeitsicher).

Konfiguration komplett über Umgebungsvariablen / GitHub-Secrets – siehe README.md.
"""

import base64
import os
import sys
import smtplib
import json
import re
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO
from zoneinfo import ZoneInfo

import requests
import openpyxl

# Lokale .env-Datei laden, falls vorhanden (für Tests auf dem eigenen Rechner).
# Auf GitHub kommen die Werte aus den Secrets – dort gibt es keine .env.
def _load_dotenv():
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            os.environ.setdefault(key.strip(), val.strip())

_load_dotenv()

# --------------------------------------------------------------------------- #
# Konfiguration                                                               #
# --------------------------------------------------------------------------- #
SHEET_ID = os.environ.get("SHEET_ID", "1IwN9I_SYDsMeZ2R7YTjYyGsdWkgI-M3BzsGRIl6bDjY")

# Datenzugriff – Variante A (empfohlen): Apps-Script-Web-App
APPS_SCRIPT_URL = os.environ.get("APPS_SCRIPT_URL", "").strip()
APPS_SCRIPT_TOKEN = os.environ.get("APPS_SCRIPT_TOKEN", "").strip()
# Datenzugriff – Variante B (Alternative): Google Service Account (JSON als String)
GOOGLE_SERVICE_ACCOUNT_JSON = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()

# Gemini
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "").strip()
GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash").strip()

# Mailversand (Gmail-SMTP mit App-Passwort – Defaults passen für Gmail/Workspace)
SMTP_HOST = os.environ.get("SMTP_HOST", "smtp.gmail.com").strip()
SMTP_PORT = int(os.environ.get("SMTP_PORT", "465"))
GMAIL_ADDRESS = os.environ.get("GMAIL_ADDRESS", "").strip()          # Absender
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "").strip()  # 16-stelliges App-Passwort
RECIPIENT = os.environ.get("RECIPIENT", "malin.mann@aufinity.com").strip()

# Steuerung
TZ = ZoneInfo("Europe/Berlin")
FORCE = os.environ.get("FORCE", "").lower() in ("1", "true", "yes")   # Zeitfenster ignorieren
DRY_RUN = os.environ.get("DRY_RUN", "").lower() in ("1", "true", "yes")  # nur ausgeben, nicht senden

# Spalten (1-basiert) laut Sheet-Aufbau
COL_THEMA = 2    # B
COL_TASK = 3     # C
COL_STAND = 6    # F
COL_BLOCKER = 8  # H
COL_DOC = 10     # J
COL_AKTUELL = 11 # K

HERE = os.path.dirname(os.path.abspath(__file__))


# --------------------------------------------------------------------------- #
# Hilfsfunktionen                                                             #
# --------------------------------------------------------------------------- #
def log(msg):
    print(f"[{datetime.now(TZ).strftime('%Y-%m-%d %H:%M:%S %Z')}] {msg}", flush=True)


def within_send_window():
    """
    GitHub-Cron läuft in UTC. Damit ganzjährig genau um 17:50 Berliner Zeit
    gesendet wird (Sommer- wie Winterzeit), triggert der Workflow zu zwei
    UTC-Zeiten und wir senden nur, wenn es in Berlin tatsächlich 17 Uhr ist
    und ein Werktag (Mo–Fr).
    """
    now = datetime.now(TZ)
    if now.weekday() > 4:  # 5=Sa, 6=So
        return False, f"Wochenende ({now.strftime('%A')})"
    if now.hour != 17:
        return False, f"außerhalb Zeitfenster (Berlin {now.strftime('%H:%M')})"
    return True, "ok"


def get_xlsx_bytes():
    """Holt das Sheet als XLSX-Bytes – bevorzugt via Apps-Script, sonst Service Account."""
    if APPS_SCRIPT_URL:
        log("Hole Sheet über Apps-Script-Web-App …")
        r = requests.get(APPS_SCRIPT_URL, params={"token": APPS_SCRIPT_TOKEN}, timeout=60)
        r.raise_for_status()
        data = r.json()
        if not data.get("ok"):
            raise RuntimeError(f"Apps-Script-Fehler: {data}")
        return base64.b64decode(data["xlsx_base64"])

    if GOOGLE_SERVICE_ACCOUNT_JSON:
        log("Hole Sheet über Google Service Account (Drive-Export) …")
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        info = json.loads(GOOGLE_SERVICE_ACCOUNT_JSON)
        creds = service_account.Credentials.from_service_account_info(
            info, scopes=["https://www.googleapis.com/auth/drive.readonly"]
        )
        service = build("drive", "v3", credentials=creds)
        req = service.files().export_media(
            fileId=SHEET_ID,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        buf = BytesIO()
        from googleapiclient.http import MediaIoBaseDownload
        downloader = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        return buf.getvalue()

    raise RuntimeError(
        "Kein Datenzugriff konfiguriert. Setze entweder APPS_SCRIPT_URL "
        "(+APPS_SCRIPT_TOKEN) oder GOOGLE_SERVICE_ACCOUNT_JSON."
    )


def _cell(ws, row, col):
    c = ws.cell(row=row, column=col)
    val = c.value
    val = str(val).strip() if val is not None else ""
    link = c.hyperlink.target if c.hyperlink else None
    return val, link


def find_header_row(ws):
    for r in range(1, 30):
        for col in range(1, 15):
            if str(ws.cell(row=r, column=col).value).strip() == "Thema":
                return r
    raise RuntimeError("Kopfzeile mit 'Thema' nicht gefunden.")


def parse_active_rows(xlsx_bytes):
    """Liest alle Zeilen mit Spalte K == 'Y' und gibt strukturierte Dicts zurück."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    header = find_header_row(ws)
    rows = []
    for r in range(header + 1, ws.max_row + 1):
        aktuell, _ = _cell(ws, r, COL_AKTUELL)
        if aktuell.upper() != "Y":
            continue
        thema, _ = _cell(ws, r, COL_THEMA)
        task, _ = _cell(ws, r, COL_TASK)
        stand, _ = _cell(ws, r, COL_STAND)
        blocker, _ = _cell(ws, r, COL_BLOCKER)
        doc_text, doc_link = _cell(ws, r, COL_DOC)
        if not (thema or task):
            continue
        rows.append({
            "thema": thema,
            "task": task,
            "stand": stand,
            "blocker": blocker,
            "doc_text": doc_text,
            "doc_link": doc_link,
        })
    return rows


def format_doc(row):
    """Doc als Markdown-Link, falls ein Link existiert – sonst reiner Text."""
    text = row["doc_text"]
    link = row["doc_link"]
    if link:
        # Rohe URL als Zelltext (z.B. GitHub/Personio) → nicht doppelt verlinken,
        # nur die URL ausgeben. Smartchips haben einen sprechenden Titel → verlinken.
        if not text or text == link or text.lower().startswith("http"):
            return link
        return f"[{text}]({link})"
    return text  # z.B. "tbd", "siehe Mail" – kein Link vorhanden


def load_style_samples():
    path = os.path.join(HERE, "style_samples.md")
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            return f.read().strip()
    return ""


def build_prompt(rows, today):
    samples = load_style_samples()

    tasks_block_lines = []
    for row in rows:
        parts = [f"- Thema: {row['thema']}"]
        parts.append(f"  Task: {row['task']}")
        if row["stand"]:
            parts.append(f"  Stand: {row['stand']}")
        if row["blocker"]:
            parts.append(f"  Blocker/Notes: {row['blocker']}")
        doc = format_doc(row)
        if doc:
            parts.append(f"  Doc: {doc}")
        tasks_block_lines.append("\n".join(parts))
    tasks_block = "\n\n".join(tasks_block_lines)

    prompt = f"""Du bist Malins Assistenz und schreibst ihr tägliches Status-Update, so wie sie es bisher an Lisa versendet hat. Schreibe auf Deutsch.

STIL – halte dich exakt an das Format der folgenden echten Beispiel-Updates von Malin:
- Beginne mit "Update {today.strftime('%-d.%-m.%y')}" und dann der Zeile "Update zu Aufgaben: Was ist heute (& in den letzten Tagen) passiert?"
- Gruppiere nach Thema. Jedes Thema ist eine eigene Überschrift-Zeile.
- Darunter kurze Stichpunkte, jeweils mit "-> " beginnend.
- Nutze knappe Status-Formulierungen wie "done", "offen:", "in progress", "scheduled", "new:", "Blocker:" – abgeleitet aus Stand und Blocker/Notes.
- Wenn ein Doc vorhanden ist, hänge den Link an den passenden Stichpunkt (Markdown-Link im Format [Titel](URL) unverändert übernehmen).
- Am Ende immer der Task-Tracker-Link: https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit
- Keine erfundenen Inhalte. Nutze nur die unten gelieferten Aufgaben. Fasse pro Thema sinnvoll zusammen, wenn mehrere Aufgaben dazu gehören.

=== BEISPIEL-UPDATES (nur als Stilvorlage, Inhalte NICHT übernehmen) ===
{samples}
=== ENDE BEISPIELE ===

Hier die aktuellen Aufgaben für das Update von heute ({today.strftime('%-d.%-m.%Y')}) – das sind alle Zeilen, die im Tracker als aktuell (Spalte K = Y) markiert sind:

{tasks_block}

Schreibe jetzt das Update. Gib NUR den fertigen Update-Text aus, ohne Vorbemerkung."""
    return prompt


def call_gemini(prompt):
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"
    headers = {"Content-Type": "application/json", "x-goog-api-key": GEMINI_API_KEY}
    body = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.4},
    }
    r = requests.post(url, headers=headers, json=body, timeout=120)
    if r.status_code != 200:
        raise RuntimeError(f"Gemini-Fehler {r.status_code}: {r.text[:500]}")
    data = r.json()
    try:
        return data["candidates"][0]["content"]["parts"][0]["text"].strip()
    except (KeyError, IndexError):
        raise RuntimeError(f"Unerwartete Gemini-Antwort: {json.dumps(data)[:500]}")


# --------------------------------------------------------------------------- #
# Mail                                                                        #
# --------------------------------------------------------------------------- #
_MD_LINK = re.compile(r"\[([^\]]+)\]\((https?://[^)]+)\)")


def to_plain(md_text):
    return _MD_LINK.sub(r"\1: \2", md_text)


def to_html(md_text):
    esc = (md_text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
    esc = _MD_LINK.sub(r'<a href="\2">\1</a>', esc)
    esc = esc.replace("\n", "<br>\n")
    return f'<div style="font-family:Arial,Helvetica,sans-serif;font-size:14px;line-height:1.5;color:#222">{esc}</div>'


def send_email(subject, md_body):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = GMAIL_ADDRESS
    msg["To"] = RECIPIENT
    msg.attach(MIMEText(to_plain(md_body), "plain", "utf-8"))
    msg.attach(MIMEText(to_html(md_body), "html", "utf-8"))

    if SMTP_PORT == 465:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=60) as server:
            server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
            server.send_message(msg)
    else:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=60) as server:
            server.starttls()
            server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
            server.send_message(msg)


# --------------------------------------------------------------------------- #
# Hauptlauf                                                                   #
# --------------------------------------------------------------------------- #
def main():
    ok, reason = within_send_window()
    if not ok and not FORCE:
        log(f"Kein Versand: {reason}. (Mit FORCE=1 erzwingbar.)")
        return 0
    if FORCE:
        log("FORCE aktiv – Zeitfenster wird ignoriert.")

    today = datetime.now(TZ)

    xlsx = get_xlsx_bytes()
    rows = parse_active_rows(xlsx)
    log(f"{len(rows)} aktuelle Aufgabe(n) (Spalte K = Y) gefunden.")
    if not rows:
        log("Keine aktuellen Aufgaben – es wird kein Update versendet.")
        return 0

    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY fehlt.")
    prompt = build_prompt(rows, today)
    update_text = call_gemini(prompt)

    subject = f"Task-Update {today.strftime('%-d.%-m.%Y')}"
    log("----- Generiertes Update -----")
    print(update_text, flush=True)
    log("------------------------------")

    if DRY_RUN:
        log("DRY_RUN aktiv – es wird keine Mail gesendet.")
        return 0

    for var, name in [(GMAIL_ADDRESS, "GMAIL_ADDRESS"), (GMAIL_APP_PASSWORD, "GMAIL_APP_PASSWORD")]:
        if not var:
            raise RuntimeError(f"{name} fehlt – Mailversand nicht möglich.")

    send_email(subject, update_text)
    log(f"Update per Mail an {RECIPIENT} gesendet.")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as e:
        log(f"FEHLER: {e}")
        sys.exit(1)
